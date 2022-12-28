import csv
import re
from collections import Counter
import openpyxl
from openpyxl.styles import Border, Side, Font
import matplotlib.pyplot as plt
import numpy as np
import time
import doctest


class Report:
    """Класс для представления отчета
    Attributes:
        file (Workbook): контейнер для всех остальных частей XLSX-документа
    """
    def __init__(self):
        """Инициализирует объект Report, создает страницу в таблице exel под именем 'Статистика по годам'
        """
        self.file = openpyxl.Workbook()

        ws = self.file.active
        ws.title = "Статистика по годам"
        self.file.create_sheet("Статистика по городам")

    def generate_excel(self, list_years, list_city):
        """Заполняет страницу 'Статистика по годам' exel таблицы переданной информацией начиная с 2007 года,
         создает и заполняет страницу "Статистика по городам" переданными данными
        Args:
            list_years([{str: int}]): список словарей содержащих статистику по годам
            list_city([{str: int}]): список словарей содержащих статистику по городам
        """
        ws = self.file["Статистика по годам"]
        ws.append(["Год", "Средняя зарплата", f"Средняя зарплата - {prof}", "Количество вакансий" , f"Количество вакансий - {prof}"])
        ws2 = self.file["Статистика по городам"]
        years = list(list_years[0].keys())
        l = len(years)
        i = 0
        flag = True
        thins = Side(border_style="thin", color="000000")
        for col in ws.iter_cols(min_row=2, max_col=5, max_row=l+1):
            min = 2007
            if flag:
                j = 0
                for cell in col:
                    cell.value = years[j]
                    cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    j += 1
                flag = False
            else:
                for cell in col:
                    cell.value = list_years[i][min]
                    cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    min += 1
                i += 1
        for col in ws.columns:
            length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = length + 2
        ws2["A1"].value = "Город"
        ws2["A1"].font = Font(bold=True)
        ws2["B1"].value = "Уровень зарплаты"
        ws2["B1"].font = Font(bold=True)
        ws2["D1"].value = "Город"
        ws2["D1"].font = Font(bold=True)
        ws2["E1"].value = "Доля вакансий"
        ws2["E1"].font = Font(bold=True)
        cities = list(list_city[0].keys())
        flag = True
        for col in ws2.iter_cols(min_row=2, max_col=2, max_row=11):
            if flag:
                j = 0
                for cell in col:
                    cell.value = cities[j]
                    cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    j += 1
                flag = False
            else:
                j = 0
                for cell in col:
                    cell.value = list_city[0][cities[j]]
                    cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    j += 1
        cities = list(list_city[1].keys())
        flag = True
        for col in ws2.iter_cols(min_row=2,min_col=4 ,max_col=5, max_row=11):
            if flag:
                j = 0
                for cell in col:
                    cell.value = cities[j]
                    cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    j += 1
                flag = False
            else:
                j = 0
                for cell in col:
                    cell.value = "{:.2%}".format(list_city[1][cities[j]])
                    cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    j += 1
        for col in ws2.columns:
            length = max(len(str(cell.value)) for cell in col)
            ws2.column_dimensions[col[0].column_letter].width = length + 2
        self.file.save('report.xlsx')


def clean_string(string):
    """очищает строку от спец символов
    Args:
        string(str): строка
    Returns:
        str: строка без спецсимволов
    """
    return ' '.join(re.sub(r"<[^>]+>", '', string).split())


def csv_reader(name):
    """очищает строку от спец символов
    Args:
        name(str): название файла
    Returns:
        list, list: списсок заголовков файла, список строк файла
    """
    csv_list = csv.reader(open(name, encoding='utf-8-sig'))
    data = [x for x in csv_list]
    return data[0], data[1::]


def csv_filer(reader):
    """очищает каждую строку файла от спец символов
        Args:
            reader(list): список строк файла
        Returns:
            list: очищенный список строк файла
        """
    all_vac = [x for x in reader if '' not in x and len(x) == len(reader[0])]
    vac = [[clean_string(y) for y in x] for x in all_vac]
    return vac


def set_graf1(salaries, prof_salaries, prof):
    """заполняет гистограмму уровня зарплат по годам
        Args:
            salaries({str: int}): словарь средней зарплаты среди всех профессии по годам
            prof_salaries({str: int}): словарь средней зарплаты для определенной профессии по годам
            prof(str): название профессии
    """
    x_axis = np.arange(len(salaries))
    years = []
    arr = []
    prof_arr = []
    for key in salaries:
        years.append(key)
        arr.append(salaries[key])
        prof_arr.append(prof_salaries[key])
    graf1.bar(x_axis - 0.2, arr, width=0.4, label='средняя з/п')
    graf1.bar(x_axis + 0.2, prof_arr, width=0.4, label='з/п ' + prof)
    graf1.set_xticks(x_axis, years, rotation='vertical')
    graf1.legend()
    graf1.set_title("Уровень заплат по годам")


def set_graf2(vacancies, prof_vacancies, prof):
    """заполняет гистограмму количества вакансий по годам
            Args:
                vacancies({str: int}): словарь среднго количества вакансий среди всех профессии по годам
                prof_vacancies({str: int}): словарь среднго количества вакансий для определенной профессии по годам
                prof(str): название профессии
        """
    x_axis = np.arange(len(vacancies))
    years = []
    arr = []
    prof_arr = []
    for key in vacancies:
        years.append(key)
        arr.append(vacancies[key])
        prof_arr.append(prof_vacancies[key])
    graf2.bar(x_axis - 0.2, arr, width=0.4, label='количество вакансий')
    graf2.bar(x_axis + 0.2, prof_arr, width=0.4, label='количество вакансий ' + prof)
    graf2.set_xticks(x_axis, years, rotation='vertical')
    graf2.legend()
    graf2.set_title("Количество вакансий по годам")


def set_graf3(salaries):
    """заполняет линейчатую диаграмму уровня зарплат по городам по убыванию
         Args:
             salaries({str: int}): словарь средней зарплаты по городам(10 самых высоких средних зарплат)
    """
    arr = []
    labels = []
    for key in salaries:
        arr.append(salaries[key])
        labels.append(key)
    graf3.barh(labels, arr)
    graf3.set_title("Уровень зарплат по городам")
    graf3.invert_yaxis()


def set_graf4(most):
    """заполняет круговую диаграмму количества вакансий по городам(10 самых высоких + все остальные)
        Args:
            most({str: int}): словарь количества вакансий по городам
    """
    arr = []
    labels = []
    s = 0
    for key in most:
        arr.append(most[key])
        s += most[key]
        labels.append(key)
    arr.append(1 - s)
    labels.append("Другие")
    graf4.pie(arr, labels=labels)
    graf4.set_title("Доля вакансий по городам")


currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


output_type = input("Какой вывод требуется(Вакансии/Статистика: ")
name = input('Введите название файла: ')
prof = input('Введите название профессии: ')
start_time = time.time()
header, vac = csv_reader(name)
vac = csv_filer(vac)
dict_naming = {}
for i in range(len(header)):
    dict_naming[header[i]] = i
salary_dynamic = {}
count_dynamic = {}
salary_prof_dynamic = {}
city_count = {}
salary_city = {}
prof_count = {}
for item in vac:
    year = int(item[dict_naming['published_at']].split('-')[0])
    if year not in count_dynamic:
        count_dynamic[year] = 0
    count_dynamic[year] += 1
    for i in range(len(item)):
        if header[i] == 'salary_from':
            salary = (float(item[i]) + float(item[i + 1])) / 2
            if item[dict_naming['salary_currency']] != 'RUR':
                salary *= currency_to_rub[item[dict_naming['salary_currency']]]
            if year not in salary_dynamic:
                salary_dynamic[year] = []
            salary_dynamic[year].append(int(salary))
            if year not in salary_prof_dynamic:
                salary_prof_dynamic[year] = []
            if prof in item[0]: salary_prof_dynamic[year].append(int(salary))
            if year not in prof_count:
                prof_count[year] = 0
            if prof in item[0]: prof_count[year] += 1
        city = item[dict_naming['area_name']]
        if city not in city_count:
            city_count[city] = 0
        city_count[city] += 1
for item in vac:
    for i in range(len(item)):
        if header[i] == 'salary_from':
            salary = (float(item[i]) + float(item[i + 1])) / 2
            city = item[dict_naming['area_name']]
            if item[dict_naming['salary_currency']] != 'RUR':
                salary *= currency_to_rub[item[dict_naming['salary_currency']]]
            if city_count[city] >= int(sum(city_count.values()) * 0.01):
                if city not in salary_city:
                    salary_city[city] = []
                salary_city[city].append(int(salary))
for key in salary_dynamic:
    salary_dynamic[key] = sum(salary_dynamic[key]) // len(salary_dynamic[key])
for key in salary_prof_dynamic:
    salary_prof_dynamic[key] = sum(salary_prof_dynamic[key]) // max(len(salary_prof_dynamic[key]), 1)
for key in salary_city:
    salary_city[key] = sum(salary_city[key]) // len(salary_city[key])
print('Динамика уровня зарплат по годам:', salary_dynamic)
print('Динамика количества вакансий по годам:', count_dynamic)
print('Динамика уровня зарплат по годам для выбранной профессии:', salary_prof_dynamic)
print('Динамика количества вакансий по годам для выбранной профессии:', prof_count)
print('Уровень зарплат по городам (в порядке убывания):', dict(Counter(salary_city).most_common(10)))
most = dict(
    Counter({k: float('{:.4f}'.format(v / sum(city_count.values()))) for k, v in city_count.items()}).most_common(10))
most = {k: v for k, v in most.items() if v >= 0.01}
print('Доля вакансий по городам (в порядке убывания):', most)

if output_type == "Вакансии":
    report = Report()
    report.generate_excel([salary_dynamic, salary_prof_dynamic, count_dynamic, prof_count], [salary_city, most])
else:
    figure, axs = plt.subplots(nrows=2, ncols=2, figsize=(10, 10), dpi=100)
    graf1 = axs[0, 0]
    graf2 = axs[0, 1]
    graf3 = axs[1, 0]
    graf4 = axs[1, 1]
    set_graf1(salary_dynamic, salary_prof_dynamic, prof)
    set_graf2(count_dynamic, prof_count, prof)
    set_graf3(dict(Counter(salary_city).most_common(10)))
    set_graf4(most)
    plt.tight_layout()
    plt.savefig("graph.png")

print("Время работы: %s seconds" % round(time.time() - start_time, 4))